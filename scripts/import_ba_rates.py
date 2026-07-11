"""Parse BA Rates.xlsx and output benchmark JSON files to data/benchmarks/."""
import json
import os
import openpyxl

def _f(v):
    """Safely convert a cell value to float; return None if not possible."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

SRC = r"C:\Users\Viknesh\Downloads\BA Rates.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "benchmarks")

wb = openpyxl.load_workbook(SRC, data_only=True)

# ── 1. BA Pincode Rates ──────────────────────────────────────────────
ws = wb["BA Rates "]
ba_pincodes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    sr, pincode, mapping, taluk, dist, state, zone, oda, airport, provider, remark, rate = row
    if pincode is None:
        continue
    ba_pincodes.append({
        "pincode": str(int(pincode)),
        "mapping_code": str(mapping).strip() if mapping else "",
        "taluk": str(taluk).strip() if taluk else "",
        "district": str(dist).strip() if dist else "",
        "state": str(state).strip() if state else "",
        "zone": str(zone).strip() if zone else "",
        "oda": str(oda).strip() if oda else "",
        "airport": str(airport).strip() if airport else "",
        "provider": str(provider).strip() if provider else "",
        "base_rate": _f(rate) or 0,
    })
with open(os.path.join(OUT, "ba_pincode_rates.json"), "w") as f:
    json.dump(ba_pincodes, f, indent=2)
print(f"ba_pincode_rates.json: {len(ba_pincodes)} rows")

# ── 2. Region Matrix ─────────────────────────────────────────────────
ws = wb["Sheet1"]
region_headers = []
region_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    vals = [v for v in row]
    if vals[0] and str(vals[0]).strip().lower() == "region":
        region_headers = [str(v).strip() if v else "" for v in vals[1:]]
        continue
    if vals[0] and str(vals[0]).strip() in ("North", "East", "West", "South", "N-East"):
        origin = str(vals[0]).strip()
        for j, v in enumerate(vals[1:]):
            if j < len(region_headers) and v is not None and region_headers[j]:
                region_rows.append({
                    "origin_region": origin,
                    "dest_region": region_headers[j],
                    "rate_per_kg": float(v) if v else 0,
                })
with open(os.path.join(OUT, "ba_region_matrix.json"), "w") as f:
    json.dump(region_rows, f, indent=2)
print(f"ba_region_matrix.json: {len(region_rows)} rows")

# ── 3. SpiceXpress Console Rates (EDS sheet) ─────────────────────────
ws = wb["EDS "]
spice_rates = []
for row in ws.iter_rows(min_row=7, values_only=True):
    origin, dest, flight, dep, arr, slab, consol, airline_rate, flat_rate, misc, outgoing, retrieval, surcharge, net_rate, console = row
    if origin is None or str(origin).strip() == "":
        continue
    spice_rates.append({
        "origin": str(origin).strip(),
        "dest": str(dest).strip() if dest else "",
        "flight": str(flight).strip() if flight else "",
        "time_slab": str(slab).strip() if slab else "",
        "consol": str(consol).strip() if consol else "",
        "airline_rate": _f(airline_rate),
        "flat_rate": _f(flat_rate),
        "misc_charges": _f(misc) or 0,
        "outgoing_charges": _f(outgoing) or 0,
        "retrieval_charges": _f(retrieval) or 0,
        "surcharge": _f(surcharge) or 0,
        "net_rate": _f(net_rate),
    })
with open(os.path.join(OUT, "spicexpress_console_rates.json"), "w") as f:
    json.dump(spice_rates, f, indent=2)
print(f"spicexpress_console_rates.json: {len(spice_rates)} rows")

# ── 4. IndiGo Console Rates ──────────────────────────────────────────
ws = wb["Console"]
indigo_rates = []
for row in ws.iter_rows(min_row=3, values_only=True):
    origin, dest, flight, dep, arr, slab, consol, airline_rate, flat_rate, misc, outgoing, retrieval, surcharge, net_rate, _, _ = row
    if origin is None or str(origin).strip() == "":
        continue
    indigo_rates.append({
        "origin": str(origin).strip(),
        "dest": str(dest).strip() if dest else "",
        "flight": str(flight).strip() if flight else "",
        "time_slab": str(slab).strip() if slab else "",
        "consol": str(consol).strip() if consol else "",
        "airline_rate": _f(airline_rate),
        "flat_rate": _f(flat_rate),
        "misc_charges": _f(misc) or 0,
        "outgoing_charges": _f(outgoing) or 0,
        "retrieval_charges": _f(retrieval) or 0,
        "surcharge": _f(surcharge) or 0,
        "net_rate": _f(net_rate),
    })
with open(os.path.join(OUT, "indigo_console_rates.json"), "w") as f:
    json.dump(indigo_rates, f, indent=2)
print(f"indigo_console_rates.json: {len(indigo_rates)} rows")

# ── 5. BOM→BLR Detailed Weight-Break Rates ───────────────────────────
ws = wb["BOM TO BLR"]
bom_blr = []
for row in ws.iter_rows(min_row=2, values_only=True):
    origin, dest, flight, product, weight, per_kg, basic_frt, de_util, fsc, ssc, awb, do_fee, coms, xray_cert, utilz, xray, subtotal, gst, total, per_kg_total, tsp, gst2, total2, tsp2, total3 = row
    if weight is None:
        continue
    bom_blr.append({
        "origin": str(origin).strip() if origin else "BOM",
        "dest": str(dest).strip() if dest else "BLR",
        "flight": str(flight).strip() if flight else "",
        "product": str(product).strip() if product else "",
        "weight_kg": float(weight),
        "per_kg_rate": _f(per_kg),
        "basic_freight": _f(basic_frt) or 0,
        "de_utilization": _f(de_util) or 0,
        "fsc": _f(fsc) or 0,
        "ssc": _f(ssc) or 0,
        "awb_fee": _f(awb) or 0,
        "do_fee": _f(do_fee) or 0,
        "coms": _f(coms) or 0,
        "xray_certification": _f(xray_cert) or 0,
        "utilization": _f(utilz) or 0,
        "xray": _f(xray) or 0,
        "subtotal": _f(subtotal) or 0,
        "gst": _f(gst) or 0,
        "total": _f(total) or 0,
        "per_kg_total": _f(per_kg_total) or 0,
    })
with open(os.path.join(OUT, "bom_blr_detailed.json"), "w") as f:
    json.dump(bom_blr, f, indent=2)
print(f"bom_blr_detailed.json: {len(bom_blr)} rows")

print("\nDone — all 5 files written to data/benchmarks/")
